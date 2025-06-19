import tkinter as tk
from tkinter import messagebox
import os
import time
from datetime import datetime
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import threading
# Selenium e Excel
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import get_column_letter

import ctypes

import shutil


# Caminho do arquivo Excel
arquivo_excel = os.path.join(os.getcwd(), "scrappingTCM.xlsx")


def executar_scraping():
    def verificar_se_planilha_esta_aberta(caminho_arquivo):
        try:
            # Tentativa de abrir o arquivo em modo exclusivo
            with open(caminho_arquivo, 'a'):
                return False  # Não está aberta
        except PermissionError:
            return True  # Está aberta em outro programa

        # Exemplo de uso no seu fluxo
    if verificar_se_planilha_esta_aberta(arquivo_excel):
        messagebox.showerror("Erro", "A planilha 'scrappingTCM.xlsx' está aberta. Por favor, feche antes de continuar.")
        return  # Interrompe a execução do programa (ou você pode usar sys.exit())

    
    hora_inicio = datetime.now()
    print(f"Hora inicial: {hora_inicio.strftime('%H:%M:%S')}")

    chrome_options = Options()
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--log-level=3")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    service = Service()
    driver = webdriver.Chrome(service=service, options=chrome_options)

    dados_coletados = []

# Lê a última sessão registrada na planilha (célula F2)
    ultima_sessao_planilha = ""
    if os.path.exists(arquivo_excel):
        wb_temp = load_workbook(arquivo_excel)
        ws_temp = wb_temp.active
        ultima_sessao_planilha = str(ws_temp["F2"].value).strip() if ws_temp["F2"].value else ""
        wb_temp.close()


    #dados_coletados

    try:
        driver.get("https://etcm.tcmrio.tc.br/processo?TipoConsulta=SessoesFechadas")
        wait = WebDriverWait(driver, 10)
        select_element = wait.until(EC.presence_of_element_located((By.ID, "SessaoIDSelecionada")))
        combo = Select(select_element)

        opcoes = []
        for opt in combo.options:
            texto = opt.text.strip()
            if "aposentadorias" in texto.lower():
                continue
            if texto == ultima_sessao_planilha:
                break  # Parar quando chegar na última já registrada
            opcoes.append((opt.get_attribute('value'), texto))

        for valor, texto in opcoes:
            wait.until(EC.presence_of_element_located((By.ID, "SessaoIDSelecionada")))
            combo = Select(driver.find_element(By.ID, "SessaoIDSelecionada"))
            combo.select_by_value(valor)

            nome_sessao = combo.first_selected_option.text.strip()

            buscar_btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@type='submit' and @value='Buscar']"))
            )
            buscar_btn.click()

            try:
                WebDriverWait(driver, 10).until(
                    lambda d: (
                        d.find_elements(By.CSS_SELECTOR, "table.table tbody tr")
                        or "nenhum processo encontrado" in d.page_source.lower()
                    )
                )

                total_linhas = len(driver.find_elements(By.CSS_SELECTOR, "table.table tbody tr")) - 1

                for i in range(1, total_linhas + 1):
                    try:
                        linha = driver.find_elements(By.CSS_SELECTOR, "table.table tbody tr")[i]
                        colunas = linha.find_elements(By.TAG_NAME, "td")
                        if len(colunas) < 5:
                            continue

                        link_elem = colunas[2].find_element(By.TAG_NAME, "a")
                        link_texto = link_elem.text.strip()
                        print(f"Coletando processo: {link_texto} | Sessão: {nome_sessao}")

                        link_href = link_elem.get_attribute("href")

                        interessado_objeto = colunas[3].text.strip()
                        interessado_lower = interessado_objeto.lower()
                        if "aposentadoria" in interessado_lower or "pensão" in interessado_lower:
                            continue

                        orgao_origem = colunas[4].text.strip()

                        driver.get(link_href)

                        teor_info = ""
                        try:
                            msg = driver.find_element(
                                By.XPATH, "//*[contains(text(), 'Inteiro Teor do processo não disponível')]"
                            )
                            teor_info = msg.text.strip()
                        except NoSuchElementException:
                            try:
                                link_inteiro_teor = WebDriverWait(driver, 3).until(
                                    EC.presence_of_element_located(
                                        (By.XPATH, "//a[contains(text(), 'Consultar inteiro teor')]")
                                    )
                                )
                                teor_info = link_inteiro_teor.get_attribute("href")
                            except TimeoutException:
                                teor_info = "Informação não encontrada."

                        data_sessoes = ""
                        try:
                            resumo = driver.find_element(By.CSS_SELECTOR, "table.table-sm")
                            header_cells = resumo.find_elements(By.TAG_NAME, "th")
                            index_sessoes = next(
                                (i for i, th in enumerate(header_cells) if th.text.strip().upper() == "DATA DAS SESSÕES"),
                                None
                            )
                            if index_sessoes is not None:
                                primeira_linha = resumo.find_element(By.CSS_SELECTOR, "tbody tr")
                                tds = primeira_linha.find_elements(By.TAG_NAME, "td")
                                if len(tds) > index_sessoes:
                                    data_sessoes = tds[index_sessoes].text.strip()
                        except NoSuchElementException:
                            data_sessoes = ""

                        todas_decisoes = []
                        try:
                            tabela_decisoes = driver.find_element(
                                By.XPATH,
                                "//h5[normalize-space(text())='Decisões do Processo']/ancestor::div[@class='row']/following-sibling::div[1]//table"
                            )
                            linhas_decisoes = tabela_decisoes.find_elements(By.CSS_SELECTOR, "tbody tr")
                            for linha_dec in linhas_decisoes:
                                colunas_dec = linha_dec.find_elements(By.TAG_NAME, "td")
                                if len(colunas_dec) >= 2:
                                    data = colunas_dec[0].text.strip()
                                    decisao = colunas_dec[1].text.strip()
                                    todas_decisoes.append((data, decisao))
                        except Exception as e:
                            print(f"Erro ao capturar decisões: {e}")

                        try:
                            div_objeto = driver.find_element(By.XPATH, "//label[@for='Processo_Objeto']/parent::div")
                            objeto_texto = div_objeto.find_element(By.TAG_NAME, "p").text.strip()
                        except NoSuchElementException:
                            objeto_texto = ""

                        ultimo_orgao = ""
                        data_recebimento = ""
                        try:
                            tabela_carga = driver.find_element(By.XPATH, "//h5[normalize-space(text())='Última Carga do Processo']/following::table[1]")
                            primeira_linha = tabela_carga.find_element(By.CSS_SELECTOR, "tbody tr")
                            tds_carga = primeira_linha.find_elements(By.TAG_NAME, "td")
                            if len(tds_carga) >= 2:
                                ultimo_orgao = tds_carga[0].text.strip()
                                data_recebimento = tds_carga[1].text.strip()
                        except NoSuchElementException:
                            pass

                        partes_procuradores = ""
                        try:
                            tabela_partes = driver.find_element(By.XPATH, "//h5[normalize-space(text())='Partes e Procuradores']/following::table[1]")
                            linhas_partes = tabela_partes.find_elements(By.TAG_NAME, "tr")
                            lista_partes = []
                            for linha_partes in linhas_partes[1:]:  # pular cabeçalho
                                colunas_partes = linha_partes.find_elements(By.TAG_NAME, "td")
                                if len(colunas_partes) == 2:
                                    tipo = colunas_partes[0].text.strip()
                                    nome = colunas_partes[1].text.strip()
                                    lista_partes.append(f"{tipo}: {nome}")
                            partes_procuradores = " ; ".join(lista_partes)
                        except NoSuchElementException:
                            partes_procuradores = ""

                        dados_coletados.append({
                            "PROCESSO": link_texto,
                            "SESSÃO MAIS RECENTES": nome_sessao,
                            "DATA SESSÃO": texto.split("-")[-1].strip() if "-" in texto else "",
                            "LINK INTEIRO TEOR": teor_info,
                            "ÓRGÃO/ENTIDADE": orgao_origem,
                            "DATAS DAS DECISÕES": "; ".join([d for d, _ in todas_decisoes]),
                            "DECISÕES": "\n".join([c for _, c in todas_decisoes]),
                            "OBJETO": objeto_texto,
                            "ÚLTIMO ÓRGÃO TRAMITADO": ultimo_orgao,
                            "DATA ÚLTIMA TRAMITAÇÃO": data_recebimento,
                            "PARTES/PROCURADORES": partes_procuradores,
                        })

                        driver.back()
                        WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "table.table tbody tr"))
                        )

                    except Exception as e:
                        print(f"Erro linha {i}: {e}")
                        continue

            except TimeoutException:
                pass

            driver.back()
            time.sleep(1)

    finally:
        driver.quit()
    if not dados_coletados:
        print("Nenhum novo dado encontrado. A planilha já está atualizada.")
        messagebox.showinfo("TCM    ", "Nenhum novo dado foi encontrado.\nA planilha já está atualizada.")
        return  # Encerra a função antes de alterar a planilha
    # --- Definição da função ---

    


    if os.path.exists(arquivo_excel):
        wb = load_workbook(arquivo_excel)
        ws = wb.active
        ultima_sessao_planilha = ws.cell(row=ws.max_row, column=6).value
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Scraping TCM"
    cabecalhos = [
        "ÓRGÃO/ENTIDADE",
        "PROCESSO",
        "LINK INTEIRO TEOR",
        "OBJETO",
        "PARTES/PROCURADORES",
        "SESSÃO MAIS RECENTES",
        "DATAS DAS DECISÕES",
        "DECISÕES COMPLETAS",
        "ÚLTIMO ÓRGÃO TRAMITADO",
        "DATA ÚLTIMA TRAMITAÇÃO",
        "AÇÕES SUGERIDAS",
        "DATA DA AÇÃO"
    ]

    # Só adiciona o cabeçalho se a planilha for nova
    if not os.path.exists(arquivo_excel):
        ws.append(cabecalhos)


    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="9cc2e5", end_color="9cc2e5", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col_num, header in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Número de linhas que vamos inserir (quantidade de registros considerando decisões múltiplas)
    num_linhas = len(dados_coletados)

    # Insere linhas em branco na posição 2 para abrir espaço
    ws.insert_rows(2, amount=num_linhas)

    linha_atual = 2
    for linha in dados_coletados:
        decisoes = linha["DECISÕES"].split("\n")
        datas = linha["DATAS DAS DECISÕES"].split("; ")

        # Monta as decisões com quebra de linha
        texto_formatado = "\n".join([f"{i + 1} - {d}" for i, d in enumerate(decisoes)])

        ws.cell(row=linha_atual, column=1, value=linha["ÓRGÃO/ENTIDADE"])
        ws.cell(row=linha_atual, column=2, value=linha["PROCESSO"])
        ws.cell(row=linha_atual, column=3, value=linha["LINK INTEIRO TEOR"])
        ws.cell(row=linha_atual, column=4, value=linha["OBJETO"])
        ws.cell(row=linha_atual, column=5, value=linha["PARTES/PROCURADORES"])
        ws.cell(row=linha_atual, column=6, value=linha["SESSÃO"])
        # Formata as datas da mesma forma que as decisões
        datas_formatadas = "\n".join([f"{i + 1} - {data}" for i, data in enumerate(datas)])
        cell_datas = ws.cell(row=linha_atual, column=7, value=datas_formatadas)
        cell_datas.alignment = Alignment(wrap_text=True)  # Permite quebra de linha

        # Insere múltiplas decisões com quebra de linha na mesma célula
        cell_decisoes = ws.cell(row=linha_atual, column=8, value=texto_formatado)
        cell_decisoes.alignment = Alignment(wrap_text=True)  # Ativa quebra de linha automática

        ws.cell(row=linha_atual, column=9, value=linha["ÚLTIMO ÓRGÃO TRAMITADO"])
        ws.cell(row=linha_atual, column=10, value=linha["DATA ÚLTIMA TRAMITAÇÃO"])

        linha_atual += 1

    wb.save(arquivo_excel)

    ultima_linha = ws.max_row
    for col in range(1, 11):
        cell = ws.cell(row=ultima_linha, column=col)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for i, column_width in enumerate([20, 20, 50, 60, 60, 15, 30, 150, 20, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = column_width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        max_lines = 1
        for cell in row:
            if cell.value:
                lines = str(cell.value).count('\n') + 1
                if lines > max_lines:
                    max_lines = lines
        # Aproximação: cada linha tem cerca de 15 de altura no Excel
        ws.row_dimensions[cell.row].height = max_lines * 15
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row):
        col_letter = get_column_letter(col[0].column)

        if col_letter in ['E','G', 'H']:
            continue  # Pula as colunas G e H

        max_length = 0
        for cell in col:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length

        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width



    def remover_processos_antigos(wb, ws):
   
        processo_col = 2  # Coluna PROCESSO (B)
        sessao_col = 6    # Coluna SESSÃO MAIS RECENTES (F)

        processos_vistos = {}
        linhas_para_remover = []

        # Varre as linhas de baixo para cima (para remover linhas sem alterar índices antes)
        for row in range(ws.max_row, 1, -1):
            processo_valor = ws.cell(row=row, column=processo_col).value
            sessao_valor = ws.cell(row=row, column=sessao_col).value

            if processo_valor is None:
                continue

            if processo_valor in processos_vistos:
                # Se já viu esse processo, marcar esta linha para remoção (linha antiga)
                linhas_para_remover.append(row)
            else:
                # Marca o processo como visto
                processos_vistos[processo_valor] = (row, sessao_valor)

        # Remove as linhas antigas marcadas
        for linha in sorted(linhas_para_remover, reverse=True):
            ws.delete_rows(linha, 1)
            
    remover_processos_antigos(wb, ws)


       
    wb.save(arquivo_excel)

    def criar_backup_oculto(nome_arquivo_origem):
            desktop_path = os.path.join(os.path.expanduser("~"), "Documents")
            data_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_backup = f"scrappingTCM_backup_{data_hora}.xlsx"
            caminho_backup = os.path.join(desktop_path, nome_backup)
            shutil.copy2(nome_arquivo_origem, caminho_backup)
            print(f"Backup criado na Área de Trabalho: {caminho_backup}")
            
            FILE_ATTRIBUTE_HIDDEN = 0x02
            try:
                ret = ctypes.windll.kernel32.SetFileAttributesW(caminho_backup, FILE_ATTRIBUTE_HIDDEN)
                if ret:
                    print("Arquivo definido como oculto.")
                else:
                    print("Falha ao definir arquivo como oculto.")
            except Exception as e:
                print(f"Erro ao definir arquivo oculto: {e}")

            return caminho_backup

    
    if os.path.exists(arquivo_excel):
            criar_backup_oculto(arquivo_excel)
    


    hora_fim = datetime.now()
    duracao = hora_fim - hora_inicio
    print(f"Hora final: {hora_fim.strftime('%H:%M:%S')}")
    print(f"Duração: {duracao}")

    messagebox.showinfo("Scraping", f"Scraping finalizado!\nTotal de registros: {len(dados_coletados)}")

def iniciar_interface():
    root = tk.Tk()
    root.title("AIT CGM")
    root.geometry("350x150")
    root.resizable(False, False)

    label = tk.Label(root, text="Carregando...", font=("Arial", 11))
    label.pack(expand=True)

    def tarefa_scraping():
        executar_scraping()
        root.quit()  # Fecha a janela após terminar

    # Rodar scraping em thread para não travar a GUI
    threading.Thread(target=tarefa_scraping, daemon=True).start()

    root.mainloop()

if __name__ == "__main__":
    iniciar_interface()